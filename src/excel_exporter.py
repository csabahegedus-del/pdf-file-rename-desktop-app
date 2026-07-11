"""
excel_exporter.py – writes the processing results to an Excel workbook.
"""
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger("pdf_rename")

_HEADERS = [
    "Eredeti fájlnév",
    "Új fájlnév",
    "Szolgáltató",
    "Számlaszám",
    "Elszámolási időszak",
    "Számla típusa",
    "Megjegyzés",
    "Feldolgozás állapota",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ERROR_FILL = PatternFill("solid", fgColor="FFCCCC")
_OK_FILL = PatternFill("solid", fgColor="CCFFCC")


def export(results: list[dict], preview_dir: Path) -> Path:
    """
    Write *results* to an Excel file in *preview_dir*.

    Each result dict must contain:
        original, new_name, provider, invoice, period, bill_type,
        notes (optional), status (optional, defaults to 'OK')
    """
    preview_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = preview_dir / f"preview_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF átnevezés"

    # Headers
    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, result in enumerate(results, start=2):
        status = result.get("status", "OK")
        row_data = [
            result.get("original", ""),
            result.get("new_name", ""),
            result.get("provider", ""),
            result.get("invoice", ""),
            result.get("period", ""),
            result.get("bill_type", ""),
            result.get("notes", ""),
            status,
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True)
        if status != "OK":
            for col in range(1, len(_HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = _ERROR_FILL
        else:
            ws.cell(row=row_idx, column=len(_HEADERS)).fill = _OK_FILL

    # Auto-fit column widths (approximation)
    col_widths = [len(h) for h in _HEADERS]
    for result in results:
        row_data = [
            result.get("original", ""),
            result.get("new_name", ""),
            result.get("provider", ""),
            result.get("invoice", ""),
            result.get("period", ""),
            result.get("bill_type", ""),
            result.get("notes", ""),
            result.get("status", "OK"),
        ]
        for i, v in enumerate(row_data):
            col_widths[i] = max(col_widths[i], min(len(str(v)), 60))
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width + 4

    ws.freeze_panes = "A2"
    wb.save(out_path)
    logger.info("Excel preview saved → %s", out_path)
    return out_path
